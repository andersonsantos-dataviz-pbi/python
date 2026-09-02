# ============================================================
# Script: Inventário de Dataflows e Status de Acesso às Datasources
# via Power BI REST API
#
# Características principais:
#   - Status calculado individualmente por Dataflow.
#   - Renovação preventiva do token a cada 50 Workspaces.
#   - Renovação preventiva do token a cada 200 Dataflows.
#   - Renovação automática e repetição da chamada quando ocorrer HTTP 401.
#   - Erros técnicos não são classificados automaticamente como "Sem acesso".
#   - Exportação com colunas de diagnóstico da API.
#
# Requisitos:
#   pip install pandas requests openpyxl
# ============================================================

from __future__ import annotations

import logging
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import requests
from requests import Response, Session


# ============================================================
# Configurações principais
# ============================================================

# Credenciais do App Registration.
# Recomendação: em produção, carregar por variáveis de ambiente ou Key Vault.
tenant_id = "81a2db7b-b536-4905-a47b-6d892bd2f210"
cliente_id = "6e486331-016d-4ac6-b2c7-9566a65a0513"
cliente_secret = "dTI8Q~8yHSt1zrYIXNIKjnUxpwO_swvU31b8VdwK"

SCOPE = "https://analysis.windows.net/powerbi/api/.default"
TOKEN_URL = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
POWERBI_API_BASE_URL = "https://api.powerbi.com/v1.0/myorg"
GROUPS_URL = f"{POWERBI_API_BASE_URL}/groups"

PAGE_SIZE = 5000
MAX_RETRIES = 6
BASE_SLEEP_SECONDS = 2
REQUEST_INTERVAL_SECONDS = 0.25
REQUEST_TIMEOUT_SECONDS = 120
TOKEN_TIMEOUT_SECONDS = 60

# Renovação preventiva por quantidade processada.
TOKEN_REFRESH_WORKSPACES = 50
TOKEN_REFRESH_DATAFLOWS = 200

OUTPUT_DIR = Path(r"D:\Dataside\RD Saúde")
OUTPUT_FILE = OUTPUT_DIR / "PowerBI_Dataflows_Status_Datasources.xlsx"
OUTPUT_SHEET_NAME = "Dataflows_Status"
OUTPUT_TABLE_NAME = "Listagem_Dataflows_Status"


# ============================================================
# Configuração de log
# ============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)


# ============================================================
# Tipos e gerenciador de token
# ============================================================

@dataclass(frozen=True)
class ApiResult:
    status: str
    http_status: Optional[int]
    error_code: str
    error_message: str


class AccessTokenManager:
    """Mantém e renova o token usado nas chamadas da Power BI REST API."""

    def __init__(self, session: Session) -> None:
        self.session = session
        self._access_token = ""
        self.refresh_count = 0

    @property
    def access_token(self) -> str:
        if not self._access_token:
            return self.refresh(reason="token inicial")
        return self._access_token

    def authorization_headers(self) -> Dict[str, str]:
        return {
            "Authorization": f"Bearer {self.access_token}",
            "Accept": "application/json",
        }

    def refresh(self, reason: str) -> str:
        """Obtém um novo token pelo fluxo OAuth client_credentials."""
        if not safe_str(tenant_id):
            raise ValueError("O parâmetro tenant_id está vazio.")
        if not safe_str(cliente_id):
            raise ValueError("O parâmetro cliente_id está vazio.")
        if not safe_str(cliente_secret):
            raise ValueError("O parâmetro cliente_secret está vazio.")

        payload = {
            "grant_type": "client_credentials",
            "client_id": safe_str(cliente_id),
            "client_secret": safe_str(cliente_secret),
            "scope": SCOPE,
        }
        headers = {
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "application/json",
        }

        logging.info("Solicitando novo token. Motivo: %s.", reason)

        response = self.session.post(
            TOKEN_URL,
            data=payload,
            headers=headers,
            timeout=TOKEN_TIMEOUT_SECONDS,
        )

        if response.status_code != 200:
            raise RuntimeError(
                "Falha ao obter token de acesso. "
                f"Status: {response.status_code}. "
                f"Resposta: {response.text}"
            )

        token_response = parse_response_json(response)
        token = safe_str(token_response.get("access_token"))

        if not token:
            raise RuntimeError(
                "Token de acesso não encontrado na resposta do Microsoft Entra ID."
            )

        self._access_token = token
        self.refresh_count += 1
        logging.info("Token renovado com sucesso. Renovação nº %s.", self.refresh_count)
        return token


# ============================================================
# Funções auxiliares
# ============================================================

def safe_str(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_response_json(response: Response) -> Dict[str, Any]:
    try:
        payload = response.json()
        return payload if isinstance(payload, dict) else {"raw": payload}
    except Exception:
        return {"raw": response.text}


def extract_api_error(response: Response) -> Tuple[str, str]:
    """Retorna código e mensagem de erro da API, quando existirem."""
    payload = parse_response_json(response)
    error = payload.get("error")

    if isinstance(error, dict):
        return safe_str(error.get("code")), safe_str(error.get("message"))

    if isinstance(error, str):
        return "", safe_str(error)

    return safe_str(payload.get("code")), safe_str(payload.get("message") or payload.get("raw"))


def parse_error_message(response: Response) -> str:
    code, message = extract_api_error(response)
    if code and message:
        return f"{code}: {message}"
    return message or code or safe_str(response.text)


def is_explicit_dataflow_unauthorized(response: Response) -> bool:
    """Somente a mensagem explícita do Dataflow recebe 'Sem acesso'."""
    code, message = extract_api_error(response)
    return (
        code.casefold() == "unauthorized"
        and message.casefold() == "user is not authorized"
    )


def request_with_retry(
    session: Session,
    token_manager: AccessTokenManager,
    method: str,
    url: str,
    params: Optional[Dict[str, Any]] = None,
    return_statuses: Tuple[int, ...] = (),
    max_retries: int = MAX_RETRIES,
) -> Response:
    """
    Executa chamada autenticada com retry.

    Regras importantes:
      - Em um HTTP 401, renova o token uma vez e repete a mesma chamada.
      - 429 e 5xx recebem retry com espera.
      - Códigos presentes em return_statuses são devolvidos ao chamador
        para classificação granular.
    """
    token_refreshed_for_401 = False

    for attempt in range(1, max_retries + 1):
        response = session.request(
            method=method,
            url=url,
            headers=token_manager.authorization_headers(),
            params=params,
            timeout=REQUEST_TIMEOUT_SECONDS,
        )

        status_code = response.status_code

        if 200 <= status_code < 300:
            return response

        if status_code == 401 and not token_refreshed_for_401:
            logging.warning(
                "HTTP 401 recebido. Renovando token e repetindo a mesma chamada: %s",
                url,
            )
            token_manager.refresh(reason=f"HTTP 401 em {url}")
            token_refreshed_for_401 = True
            continue

        if status_code in return_statuses:
            return response

        if status_code == 429:
            retry_after = safe_str(response.headers.get("Retry-After"))
            sleep_seconds = (
                int(retry_after)
                if retry_after.isdigit()
                else BASE_SLEEP_SECONDS * attempt
            )
            logging.warning(
                "Rate limit 429. Tentativa %s/%s. Pausa de %s segundos.",
                attempt,
                max_retries,
                sleep_seconds,
            )
            time.sleep(sleep_seconds)
            continue

        if 500 <= status_code <= 599:
            sleep_seconds = BASE_SLEEP_SECONDS * attempt
            logging.warning(
                "Erro HTTP %s temporário. Tentativa %s/%s. Pausa de %s segundos.",
                status_code,
                attempt,
                max_retries,
                sleep_seconds,
            )
            time.sleep(sleep_seconds)
            continue

        raise RuntimeError(
            "Erro ao chamar a Power BI REST API. "
            f"Status: {status_code}. URL: {url}. "
            f"Mensagem: {parse_error_message(response)}"
        )

    raise RuntimeError(
        f"Falha após {max_retries} tentativas ao chamar a URL: {url}"
    )


# ============================================================
# Leitura e transformação de Workspaces
# ============================================================

def get_powerbi_groups(
    session: Session,
    token_manager: AccessTokenManager,
) -> List[Dict[str, Any]]:
    all_groups: List[Dict[str, Any]] = []
    skip = 0

    logging.info("Iniciando leitura da API Groups...")

    while True:
        params = {"$top": PAGE_SIZE, "$skip": skip}
        response = request_with_retry(
            session=session,
            token_manager=token_manager,
            method="GET",
            url=GROUPS_URL,
            params=params,
        )

        groups = parse_response_json(response).get("value", [])
        if not isinstance(groups, list):
            raise RuntimeError(
                "Formato inesperado na API Groups: 'value' não é uma lista."
            )

        all_groups.extend(groups)
        logging.info(
            "Workspaces recebidos nesta página: %s. Total acumulado: %s.",
            len(groups),
            len(all_groups),
        )

        if len(groups) < PAGE_SIZE:
            break

        skip += PAGE_SIZE
        time.sleep(REQUEST_INTERVAL_SECONDS)

    return all_groups


def transform_groups_to_dataframe(groups: List[Dict[str, Any]]) -> pd.DataFrame:
    final_columns = ["Workspace ID", "Workspace Name"]
    if not groups:
        return pd.DataFrame(columns=final_columns)

    df = pd.DataFrame(groups)
    for column in ("id", "name"):
        if column not in df.columns:
            df[column] = ""

    df = df[["id", "name"]].rename(
        columns={"id": "Workspace ID", "name": "Workspace Name"}
    )
    df["Workspace ID"] = df["Workspace ID"].map(safe_str)
    df["Workspace Name"] = df["Workspace Name"].map(safe_str)
    df = df[df["Workspace ID"] != ""].copy()
    df = df[
        df["Workspace Name"].str.casefold() != "admin monitoring"
    ].copy()
    df.sort_values(["Workspace Name", "Workspace ID"], inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df[final_columns]


# ============================================================
# Leitura e transformação de Dataflows por Workspace
# ============================================================

def get_dataflows_by_workspace(
    session: Session,
    token_manager: AccessTokenManager,
    workspace_id: str,
    workspace_name: str,
) -> List[Dict[str, Any]]:
    url = f"{POWERBI_API_BASE_URL}/groups/{workspace_id}/dataflows"
    response = request_with_retry(
        session=session,
        token_manager=token_manager,
        method="GET",
        url=url,
    )

    dataflows = parse_response_json(response).get("value", [])
    if not isinstance(dataflows, list):
        raise RuntimeError(
            "Formato inesperado na API Dataflows. "
            f"Workspace ID: {workspace_id}."
        )

    return [
        {
            "Workspace ID": workspace_id,
            "Workspace Name": workspace_name,
            "objectId": dataflow.get("objectId"),
            "name": dataflow.get("name"),
        }
        for dataflow in dataflows
        if isinstance(dataflow, dict)
    ]


def get_all_dataflows_from_workspaces(
    session: Session,
    token_manager: AccessTokenManager,
    df_workspaces: pd.DataFrame,
    refresh_every: int = TOKEN_REFRESH_WORKSPACES,
) -> List[Dict[str, Any]]:
    all_dataflows: List[Dict[str, Any]] = []
    total_workspaces = len(df_workspaces)

    for position, (_, row) in enumerate(df_workspaces.iterrows(), start=1):
        if (
            refresh_every > 0
            and position > 1
            and (position - 1) % refresh_every == 0
        ):
            token_manager.refresh(
                reason=f"renovação preventiva após {position - 1} Workspaces"
            )

        workspace_id = safe_str(row["Workspace ID"])
        workspace_name = safe_str(row["Workspace Name"])

        logging.info(
            "Processando Workspace %s/%s: %s | %s",
            position,
            total_workspaces,
            workspace_name,
            workspace_id,
        )

        if not workspace_id:
            logging.warning("Workspace ignorado por ausência de Workspace ID.")
            continue

        try:
            dataflows = get_dataflows_by_workspace(
                session=session,
                token_manager=token_manager,
                workspace_id=workspace_id,
                workspace_name=workspace_name,
            )
            all_dataflows.extend(dataflows)
            logging.info(
                "Dataflows encontrados no Workspace '%s': %s",
                workspace_name,
                len(dataflows),
            )
        except Exception as ex:
            # Uma falha ao listar um Workspace não cria linhas falsas para seus
            # Dataflows e não altera os Dataflows de outros Workspaces.
            logging.error(
                "Erro ao listar Dataflows do Workspace '%s' (%s). "
                "O processamento seguirá. Erro: %s",
                workspace_name,
                workspace_id,
                ex,
            )

        time.sleep(REQUEST_INTERVAL_SECONDS)

    return all_dataflows


def transform_dataflows_to_dataframe(
    dataflows: List[Dict[str, Any]],
) -> pd.DataFrame:
    final_columns = [
        "Workspace ID",
        "Workspace Name",
        "Dataflow ID",
        "Dataflow Name",
    ]
    if not dataflows:
        return pd.DataFrame(columns=final_columns)

    df = pd.DataFrame(dataflows)
    for column in ("Workspace ID", "Workspace Name", "objectId", "name"):
        if column not in df.columns:
            df[column] = ""

    df = df[["Workspace ID", "Workspace Name", "objectId", "name"]].rename(
        columns={"objectId": "Dataflow ID", "name": "Dataflow Name"}
    )

    for column in final_columns:
        df[column] = df[column].map(safe_str)

    df = df[df["Dataflow ID"] != ""].copy()
    df.drop_duplicates(
        subset=["Workspace ID", "Dataflow ID"],
        keep="first",
        inplace=True,
    )
    df.sort_values(
        ["Workspace Name", "Dataflow Name", "Dataflow ID"],
        inplace=True,
    )
    df.reset_index(drop=True, inplace=True)
    return df[final_columns]


# ============================================================
# Consulta e classificação individual por Dataflow
# ============================================================

def classify_datasource_response(response: Response) -> ApiResult:
    http_status = response.status_code
    error_code, error_message = extract_api_error(response)

    if 200 <= http_status < 300:
        return ApiResult("Normal", http_status, "", "")

    if is_explicit_dataflow_unauthorized(response):
        return ApiResult("Sem acesso", http_status, error_code, error_message)

    if http_status == 401:
        # Este 401 ocorreu mesmo após a renovação automática do token.
        return ApiResult(
            "Erro de autenticação",
            http_status,
            error_code,
            error_message or "HTTP 401 após renovação automática do token.",
        )

    if http_status == 403:
        return ApiResult(
            "Erro de permissão da API",
            http_status,
            error_code,
            error_message,
        )

    if http_status == 400:
        return ApiResult(
            "Erro do Dataflow/API",
            http_status,
            error_code,
            error_message,
        )

    return ApiResult(
        "Erro de consulta",
        http_status,
        error_code,
        error_message or parse_error_message(response),
    )


def get_dataflow_status_by_datasources_api(
    session: Session,
    token_manager: AccessTokenManager,
    workspace_id: str,
    dataflow_id: str,
) -> ApiResult:
    url = (
        f"{POWERBI_API_BASE_URL}/groups/{workspace_id}"
        f"/dataflows/{dataflow_id}/datasources"
    )

    response = request_with_retry(
        session=session,
        token_manager=token_manager,
        method="GET",
        url=url,
        return_statuses=(400, 401, 403),
    )
    return classify_datasource_response(response)


def enrich_dataflows_with_status(
    session: Session,
    token_manager: AccessTokenManager,
    df_dataflows: pd.DataFrame,
    refresh_every: int = TOKEN_REFRESH_DATAFLOWS,
) -> pd.DataFrame:
    final_columns = [
        "Workspace ID",
        "Workspace Name",
        "Dataflow ID",
        "Dataflow Name",
        "Status Dataflow",
        "HTTP Status",
        "Código Erro API",
        "Mensagem Erro API",
    ]

    if df_dataflows.empty:
        return pd.DataFrame(columns=final_columns)

    df_result = df_dataflows.copy()
    df_result["Status Dataflow"] = ""
    df_result["HTTP Status"] = pd.Series([pd.NA] * len(df_result), dtype="Int64")
    df_result["Código Erro API"] = ""
    df_result["Mensagem Erro API"] = ""

    total_rows = len(df_result)

    for position, (row_index, row) in enumerate(df_result.iterrows(), start=1):
        if (
            refresh_every > 0
            and position > 1
            and (position - 1) % refresh_every == 0
        ):
            token_manager.refresh(
                reason=f"renovação preventiva após {position - 1} Dataflows"
            )

        workspace_id = safe_str(row["Workspace ID"])
        workspace_name = safe_str(row["Workspace Name"])
        dataflow_id = safe_str(row["Dataflow ID"])
        dataflow_name = safe_str(row["Dataflow Name"])

        logging.info(
            "Consultando Dataflow %s/%s: %s | %s",
            position,
            total_rows,
            workspace_name,
            dataflow_name,
        )

        if not workspace_id or not dataflow_id:
            result = ApiResult(
                "Erro de identificação",
                None,
                "",
                "Workspace ID ou Dataflow ID ausente.",
            )
        else:
            try:
                result = get_dataflow_status_by_datasources_api(
                    session=session,
                    token_manager=token_manager,
                    workspace_id=workspace_id,
                    dataflow_id=dataflow_id,
                )
            except requests.RequestException as ex:
                result = ApiResult("Erro de conexão", None, "", safe_str(ex))
            except Exception as ex:
                result = ApiResult("Erro de consulta", None, "", safe_str(ex))

        # Atribuição feita exclusivamente na linha do Dataflow atual.
        df_result.at[row_index, "Status Dataflow"] = result.status
        df_result.at[row_index, "HTTP Status"] = result.http_status
        df_result.at[row_index, "Código Erro API"] = result.error_code
        df_result.at[row_index, "Mensagem Erro API"] = result.error_message

        time.sleep(REQUEST_INTERVAL_SECONDS)

    df_result = df_result[final_columns].copy()
    df_result.sort_values(
        ["Workspace Name", "Dataflow Name", "Dataflow ID"],
        inplace=True,
    )
    df_result.reset_index(drop=True, inplace=True)
    return df_result


# ============================================================
# Exportação para Excel
# ============================================================

def write_excel_for_powerbi(df: pd.DataFrame, output_file: Path) -> None:
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    output_file.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=OUTPUT_SHEET_NAME, index=False)
        worksheet = writer.sheets[OUTPUT_SHEET_NAME]

        max_row = worksheet.max_row
        max_col = worksheet.max_column

        if max_row >= 2 and max_col >= 1:
            table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
            table = Table(displayName=OUTPUT_TABLE_NAME, ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)

        widths = {
            "A": 42,
            "B": 70,
            "C": 42,
            "D": 70,
            "E": 28,
            "F": 15,
            "G": 30,
            "H": 100,
        }
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width

    logging.info("Arquivo Excel gerado com sucesso: %s", output_file)


# ============================================================
# Execução principal
# ============================================================

def main() -> None:
    started_at = datetime.now()

    logging.info("============================================================")
    logging.info("Iniciando inventário de Workspaces e Dataflows")
    logging.info("============================================================")

    with requests.Session() as session:
        token_manager = AccessTokenManager(session)
        token_manager.refresh(reason="início da execução")

        groups = get_powerbi_groups(session, token_manager)
        df_workspaces = transform_groups_to_dataframe(groups)
        logging.info("Workspaces armazenados em memória: %s", len(df_workspaces))

        dataflows = get_all_dataflows_from_workspaces(
            session=session,
            token_manager=token_manager,
            df_workspaces=df_workspaces,
            refresh_every=TOKEN_REFRESH_WORKSPACES,
        )
        df_dataflows = transform_dataflows_to_dataframe(dataflows)
        logging.info("Dataflows armazenados em memória: %s", len(df_dataflows))

        df_final = enrich_dataflows_with_status(
            session=session,
            token_manager=token_manager,
            df_dataflows=df_dataflows,
            refresh_every=TOKEN_REFRESH_DATAFLOWS,
        )

        write_excel_for_powerbi(df_final, OUTPUT_FILE)

    print("\nLista de Dataflows com Status de Acesso:\n")
    if df_final.empty:
        print("Nenhum Dataflow encontrado.")
    else:
        print(df_final.to_string(index=False))

    elapsed = datetime.now() - started_at
    logging.info("============================================================")
    logging.info("Resumo da execução")
    logging.info("============================================================")
    logging.info("Total de Workspaces lidos: %s", len(df_workspaces))
    logging.info("Total de Dataflows lidos: %s", len(df_dataflows))
    logging.info("Total de registros exportados: %s", len(df_final))
    logging.info("Total de tokens emitidos: %s", token_manager.refresh_count)
    logging.info("Arquivo exportado: %s", OUTPUT_FILE)
    logging.info("Tempo total de execução: %s", elapsed)


if __name__ == "__main__":
    main()
